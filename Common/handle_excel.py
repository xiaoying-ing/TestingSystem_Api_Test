# -*- coding:utf-8 -*-

'''
@Author  :   xiaoyin_ing

@Email   :   2455899418@qq.com

@Software:   PyCharm

@File    :   handle_excel.py

@Time    :   2020/7/2 10:20

@Desc    :

'''
from openpyxl import load_workbook
import json


class HandleExcel:
    def __init__(self, file_path, sheet_name):
        self.wb = load_workbook(file_path)  # 工作簿对象由load_workbook(file_path)获取
        self.sh = self.wb[sheet_name]  # 表单对象由工作簿对象[表单名称]获取

    def __red_titles(self):
        titles = []
        for items in list(self.sh.rows)[0]:
            titles.append(items.value)
        return titles                                          # 读取表单对象首行作为测试数据的key

    def read_test_cases_datas(self):
        test_case_datas = []
        titles = self.__red_titles()
        for items in list(self.sh.rows)[1:]:
            values = []
            for val in items:
                values.append(val.value)
            res = dict(zip(titles, values))
            test_case_datas.append(res)
        return test_case_datas                                # 读取完整的测试数据

    def close_file(self):
        self.wb.close()

'''
            res["params"] = json.loads(res["params"])
            res["expect_res"] = json.loads(res["expect_res"])
            res['case_name'] = res['case_name'].replace("\n", "--")
            res['number'] = str(res['number'])                               # 处理表中的数据
'''